import os
import glob
from .models import Research
import openpyxl
import json

class Data():
    def find_all_files(self):
        file_names = os.listdir('Data\\sheets')
        return file_names
    
    def delete_models(self):
        Research.objects.all().delete()

    def get_info(self, sheet, ind, level, node, parent = None):
        label = sheet.cell(row = ind, column = 1).value
        start = ind + 1
        while sheet.cell(row = ind, column = 13).value != 'head3':
            ind = ind + 1
        source = sheet.cell(row = ind, column = 1).value
        end = ind - 1
        
        component_data = []
        table_data = []
        header = []
        cagr=[]
        labels = []
        for row in sheet.iter_rows(min_row = start, max_col = 12, max_row = start, values_only=True):
            header = row
            labels = row[1:-1]

        for row in sheet.iter_rows(min_row = start + 1, max_col = 12, max_row = end, values_only=True):
            table_data.append(row)

        for row in table_data:
            component = row[0]
            component_values = row[1:-1]
            cagr_val = row[-1]
            component_data.append({
                'label': component,
                'data': component_values
            })
            cagr.append(cagr_val * 100)

        table_dict = {
            'labels': labels,
            'header': header,
            'datasets': component_data,
            'cagr': cagr
        }

        json_data = json.dumps(table_dict)

        while sheet.cell(row = ind, column = 13).value != 'head4':
            ind = ind + 1
        summary = sheet.cell(row = ind, column = 1).value

        node.label = label
        node.source = source
        node.data = json_data
        node.summary = summary
        node.parent = parent
        node.level = level
        node.save()
        return ind

    def condition(self, sheet, ind, max):
        if ind <= max and sheet.cell(row = ind, column = 13).value != 'head1' and sheet.cell(row = ind, column = 13).value != 'head5' and sheet.cell(row = ind, column = 13).value != 'head6' and sheet.cell(row = ind, column = 13).value != 'head7' and sheet.cell(row = ind, column = 13).value != 'head8' and sheet.cell(row = ind, column = 13).value != 'head9' and sheet.cell(row = ind, column = 13).value != 'head10':
            return True
        else:
            return False
    
    def add_models(self):
        files = self.find_all_files()
        for file in files:
            workbook = openpyxl.load_workbook('Data\\sheets\\' + file)
            list_of_sheets = workbook.sheetnames
            sheet_heads = workbook.active
            for num, sheet_name in enumerate(list_of_sheets[1:], 4):
                sheet = workbook[sheet_name]
                head0 = Research(title = sheet_heads.cell(row = num, column = 3).value, level = 0)
                head0.save()
                max = sheet.max_row
                ind = 0
                while ind <= max:
                    ind = ind + 1
                    while sheet.cell(row = ind, column = 13).value == 'head1':
                        title = sheet.cell(row = ind, column = 1).value
                        head1 = Research(title = title)
                        while sheet.cell(row = ind, column = 13).value != 'head2':
                            ind = ind + 1
                        ind = self.get_info(sheet, ind, 1, head1, head0)
                        while self.condition(sheet, ind, max):
                            ind = ind + 1

                        while sheet.cell(row = ind, column = 13).value == 'head5' or sheet.cell(row = ind, column = 13).value == 'head6':
                            title = sheet.cell(row = ind, column = 1).value
                            head5 = Research(title = title)
                            while sheet.cell(row = ind, column = 13).value != 'head2' and sheet.cell(row = ind, column = 13).value != 'head6':
                                ind = ind + 1
                            if sheet.cell(row = ind, column = 13).value == 'head6':
                                title = sheet.cell(row = ind, column = 1).value
                                head5.title = title
                                while sheet.cell(row = ind, column = 13).value != 'head2':
                                    ind = ind + 1

                            ind = self.get_info(sheet, ind, 2, head5, head1)

                            while self.condition(sheet, ind, max):
                                ind = ind + 1

                            while sheet.cell(row = ind, column = 13).value == 'head7' or sheet.cell(row = ind, column = 13).value == 'head8':
                                title = sheet.cell(row = ind, column = 1).value
                                head7 = Research(title = title)
                                while sheet.cell(row = ind, column = 13).value != 'head2' and sheet.cell(row = ind, column = 13).value != 'head8':
                                    ind = ind + 1
                                if sheet.cell(row = ind, column = 13).value == 'head8':
                                    title = sheet.cell(row = ind, column = 1).value
                                    head7.title = title
                                    while sheet.cell(row = ind, column = 13).value != 'head2':
                                        ind = ind + 1

                                ind = self.get_info(sheet, ind, 3, head7, head5)

                                while self.condition(sheet, ind, max):
                                    ind = ind + 1

                                while sheet.cell(row = ind, column = 13).value == 'head9' or sheet.cell(row = ind, column = 13).value == 'head10':
                                    title = sheet.cell(row = ind, column = 1).value
                                    head9 = Research(title = title)
                                    while sheet.cell(row = ind, column = 13).value != 'head2' and sheet.cell(row = ind, column = 13).value != 'head10':
                                        ind = ind + 1
                                    if sheet.cell(row = ind, column = 13).value == 'head10':
                                        title = sheet.cell(row = ind, column = 1).value
                                        head9.title = title
                                        while sheet.cell(row = ind, column = 13).value != 'head2':
                                            ind = ind + 1

                                    ind = self.get_info(sheet, ind, 4, head9, head7)

                                    while self.condition(sheet, ind, max):
                                        ind = ind + 1
                            
    def run_all_functions(self):
        self.delete_models()
        self.add_models()